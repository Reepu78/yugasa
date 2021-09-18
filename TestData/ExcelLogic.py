import openpyxl

from Config.config import ConfigData


class TestDataFromExcel:
    wb = openpyxl.load_workbook(ConfigData.TESTDATA_EXCEL_PATH)
    USERNAME = wb['LoginPage']['A2'].value
    NEW_PASSWORD = wb['LoginPage']['E2'].value

    """Login Page"""

    def readPassword(self):
        wb1 = openpyxl.load_workbook(ConfigData.TESTDATA_EXCEL_PATH)
        return wb1['LoginPage']['B2'].value

    def readIntentVariable(self):
        wb1 = openpyxl.load_workbook(ConfigData.TESTDATA_EXCEL_PATH)
        return wb1['CreateIntent']['A5'].value

    def read_addSlot_title_variable(self):
        wb1 = openpyxl.load_workbook(ConfigData.TESTDATA_EXCEL_PATH)
        return wb1['SLOTSPAGE']['A2'].value

    INVALID_USERNAME = wb['LoginPage']['C2'].value
    INVALID_PASSWORD = wb['LoginPage']['D2'].value

    INVALID_CONFIRM_PASSWORD = wb['LoginPage']['F2'].value

    """Registration Page"""
    NINE_DIGIT_MOBILE_NUMBER = wb['RegistrationPage']['A2'].value
    ELEVEN_DIGIT_MOBILE_NUMBER = wb['RegistrationPage']['B2'].value
    ALPHANUM_MOBILE_NUMBER = wb['RegistrationPage']['C2'].value
    SPECIAL_CHAR_MOBILE_NUMBER = wb['RegistrationPage']['D2'].value

    INV_PWD_LESS_THAN_EIGHT = wb['RegistrationPage']['E2'].value
    INV_PWD_NO_LWRCASE = wb['RegistrationPage']['F2'].value
    INV_PWD_NO_UPPRCASE = wb['RegistrationPage']['G2'].value
    INV_PWD_NO_SPCLCHAR = wb['RegistrationPage']['H2'].value
    INV_PWD_NO_NUMBER = wb['RegistrationPage']['I2'].value
    INV_UNAME_SPECIAL_CHAR = wb['RegistrationPage']['J2'].value

    """Edit Profile Page"""
    Edit_Company_Name = wb['EditProfilePage']['A2'].value

    def write_in_excel(sheetName, rowNum, colNum, val):
        workbookPath = ConfigData.TESTDATA_EXCEL_PATH
        wbk = openpyxl.load_workbook(workbookPath)
        wb = openpyxl.load_workbook(workbookPath)

        sh = wb[sheetName]
        sh.cell(row=rowNum, column=colNum, value=val)

        wb.save(workbookPath)
        wb.close

        # for wks in wbk.worksheets:
        #     wks.cell(row=rowNum, column=colNum).value = val

